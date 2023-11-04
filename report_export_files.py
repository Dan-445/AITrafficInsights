from typing import Dict, Any

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

# import argparse

# parser = argparse.ArgumentParser("file_data_path")
# parser.add_argument("--file_data_path", type=str, help="file_data_path")
# args = parser.parse_args()
# file_path = args.file_data_path
file_path = 'detected_vehicle_count.csv'

direction_mapping = {
    ('north', 'north'): ['Northbound', 'thru'],
    ('south', 'south'): ['Southbound', 'thru'],
    ('east', 'east'): ['Eastbound', 'thru'],
    ('west', 'west'): ['Westbound', 'thru'],
    ('south', 'west'): ['Southbound', 'left'],
    ('north', 'south'): ['Northbound', 'U-turn'],
    ('north', 'east'): ['Northbound', 'left'],
    ('north', 'west'): ['Northbound', 'right'],
    ('south', 'north'): ['Southbound', 'U-turn'],
    ('south', 'nast'): ['Southbound', 'right'],
    ('east', 'north'): ['Eastbound', 'right'],
    ('east', 'south'): ['Eastbound', 'left'],
    ('east', 'west'): ['Eastbound', 'U-turn'],
    ('west', 'north'): ['Westbound', 'left'],
    ('west', 'south'): ['Westbound', 'right'],
    ('west', 'east'): ['Westbound', 'U-turn']
}


def convert_time_format(time_str):
    time_parts = time_str.split(':')
    total_seconds = int(time_parts[0]) + float(time_parts[1])
    converted_time = (datetime.min + timedelta(seconds=total_seconds)).time()
    return converted_time.strftime("%I:%M:%S %p")


df_vehicle_detection = pd.read_csv(file_path)
df_vehicle_detection['Orientation'] = df_vehicle_detection['Orientation'].str.lower()
df_vehicle_detection['Converted_Time'] = df_vehicle_detection['time']
df_vehicle_detection.sort_values(by='Converted_Time', inplace=True)
all_car_ids = df_vehicle_detection['ID'].unique()
detection_tracking_list = []


# Function to remove rows where the Orientation values are not repeated for the next two rows
def remove_non_repeated_orientations(df_car_tracking):
    df_car_tracking.reset_index(drop=True, inplace=True)
    indices_to_drop = []
    for row in range(len(df_car_tracking) - 2):
        if not (df_car_tracking.iloc[row]["Orientation"] == df_car_tracking.iloc[row + 1]["Orientation"]):
            indices_to_drop.extend([row])
    for row in range(len(df_car_tracking) - 1, 1, -1):
        if not (df_car_tracking.iloc[row]["Orientation"] == df_car_tracking.iloc[row - 1]["Orientation"]):
            indices_to_drop.extend([row])
    df_cleaned = df_car_tracking.drop(indices_to_drop, axis=0)
    return df_cleaned


for car_id in all_car_ids:
    vehicle_tracking_output = {}
    vehicle_tracking_dataset = df_vehicle_detection[df_vehicle_detection['ID'] == car_id]
    if not vehicle_tracking_dataset.empty:
        vehicle_tracking_output['ID'] = vehicle_tracking_dataset['ID'].values[0]
        vehicle_tracking_output['Class'] = vehicle_tracking_dataset['Class'].values[0]
        start_time = vehicle_tracking_dataset['Converted_Time'].values[0]
        vehicle_tracking_output['start_time'] = start_time
        start_orientation = vehicle_tracking_dataset['Orientation'].tolist()[0]
        end_orientation = vehicle_tracking_dataset['Orientation'].tolist()[-1]
        orientation = direction_mapping.get((start_orientation, end_orientation), [start_orientation, end_orientation])
        vehicle_tracking_output['start_orientation'] = orientation[0]
        vehicle_tracking_output['end_orientation'] = orientation[1]
        detection_tracking_list.append(vehicle_tracking_output)
direction_detected_dataset = pd.DataFrame(detection_tracking_list)

pivot_table = pd.pivot_table(direction_detected_dataset, values='Class', index=['start_time'],
                             columns=['start_orientation', 'end_orientation'], aggfunc=lambda x: list(x), fill_value=0)
all_columns = [('Southbound', 'right'), ('Southbound', 'thru'), ('Southbound', 'left'), ('Southbound', 'peds'),
               ('Southbound', 'U-turn'), ('Southbound', 'Vehicle Class'), ('Westbound', 'right'), ('Westbound', 'thru'),
               ('Westbound', 'left'), ('Westbound', 'peds'), ('Westbound', 'U-turn'), ('Westbound', 'Vehicle Class'),
               ('Northbound', 'right'), ('Northbound', 'thru'), ('Northbound', 'left'), ('Northbound', 'peds'),
               ('Northbound', 'U-turn'), ('Northbound', 'Vehicle Class'), ('Eastbound', 'right'), ('Eastbound', 'thru'),
               ('Eastbound', 'left'), ('Eastbound', 'peds'), ('Eastbound', 'U-turn'), ('Eastbound', 'Vehicle Class')]
pivot_table = pivot_table.reindex(columns=all_columns, fill_value=0)
pivot_table.columns.names = ["", ""]
column_dict = {col: 0 for col in all_columns}
new_row = pd.Series(column_dict, name='total_count')
pivot_table = pd.concat([pivot_table, new_row.to_frame().T])


def map_class(row):
    all_classes = []
    for direction in ['right', 'thru', 'left', 'U-turn']:
        if isinstance(row[direction], list):
            all_classes.extend(row[direction])
            row[direction] = len(row[direction])
    row['Vehicle Class'] = ','.join(all_classes)
    if isinstance(row["peds"], list):
        row["peds"] = len(row["peds"])
    return row


for direction in ['Southbound', 'Westbound', 'Northbound', 'Eastbound']:
    direction_columns: Dict[Any, Any] = {}
    df_data = pivot_table[direction]
    df_data = df_data.apply(map_class, axis=1)
    for vehicle_direction in ['right', 'thru', 'left', 'peds', 'U-turn']:
        direction_columns[vehicle_direction] = df_data[vehicle_direction].sum()
    df_data.loc['total_count'] = direction_columns
    pivot_table[direction] = df_data

# Create a new workbook
workbook = Workbook()
worksheet = workbook.active

# Data and formatting for the first set of cells
header_data = ['Start Date:', 'Start Time:', 'Site Code:', 'Comment 1:', 'Comment 2:', 'Comment 3:', 'Comment 4:']
for row, data in enumerate(header_data, start=1):
    cell = worksheet.cell(row, 1, data)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

# Data and formatting for the second set of cells
content_data = ['', '6:00:00 AM', '', 'Default Comments', 'Change These in The Preferences Window',
                'Select File/Preference in the Main Screen', 'Then Click the Comments Tab']
for row, data in enumerate(content_data, start=1):
    cell = worksheet.cell(row, 3, data)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    worksheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=22)

# Data and formatting for the columns in rows 8 and 9
header_data = ['Southbound', 'Westbound', 'Northbound', 'Eastbound']
for col, data in enumerate(header_data, start=1):
    col += (col - 1) * 5
    cell = worksheet.cell(8, col + 1, data)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Gray color fill
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    worksheet.merge_cells(start_row=8, start_column=col + 1, end_row=8, end_column=col + 4)
    additional_cell = worksheet.cell(8, col + 5, 'Additional')
    movement_cell = worksheet.cell(8, col + 6, 'Movement')

    for cell in [additional_cell, movement_cell]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Gray color fill
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

header_data = ['Right', 'Thru', 'Left', 'Peds', 'U-Turns', 'Vehicle Class'] * 4
cell = worksheet.cell(9, 1, 'Start Time')
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Gray color fill
cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))

for col, data in enumerate(header_data, start=1):
    cell = worksheet.cell(9, col + 1, data)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Gray color fill
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

# Adjust row heights
worksheet.row_dimensions[8].height = 30
worksheet.row_dimensions[9].height = 30

# Convert the pivot table to a list of lists
pivot_values = pivot_table.reset_index().values.tolist()
# Write the script and the pivot table to the same sheet
for row_data, row in enumerate(pivot_values, start=1):
    for column_data, value in enumerate(row, start=1):
        worksheet.cell(row=row_data + len(content_data) + 2, column=column_data, value=value)

# Save the workbook
workbook.save('vehicle_counting_export.xlsx')
print("Pivot table have been written to the same Excel sheet successfully")
